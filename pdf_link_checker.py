"""
PDF Link Checker  —  Ultra-Fast Edition
========================================
Validates ALL email and web addresses in PDFs with NO page limit.

Speed vs original:
  • PyMuPDF (fitz) replaces pdfplumber  →  10–50x faster text extraction
  • Links + text extracted in ONE pass per page  →  single file open per worker
  • ProcessPoolExecutor chunks pages across all CPU cores  →  true parallelism
  • Spatial Y-bucket index  →  O(1) link lookup instead of O(n) scan

Requirements:
    pip install pymupdf openpyxl
    (pymupdf replaces pdfplumber AND pypdf — much faster)

Usage:
    python pdf_link_checker.py                   # GUI file picker
    python pdf_link_checker.py document.pdf
    python pdf_link_checker.py document.pdf -o report.xlsx -v
"""

import re, os, sys, argparse, math
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

# ── Regex (compiled once) ────────────────────────────────────────────────────
EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
URL_REGEX   = re.compile(r"(?:https?://|www\.)[a-zA-Z0-9\-._~:/?#\[\]@!$&'()*+,;=%]+")


# ── Data classes ─────────────────────────────────────────────────────────────
@dataclass
class LinkAnnotation:
    page_num: int
    uri: str
    rect: tuple          # (x0, y0, x1, y1)

@dataclass
class TextToken:
    page_num: int
    text: str
    x0: float; y0: float; x1: float; y1: float

@dataclass
class ExcelRow:
    pdf_link: str
    hyperlink_text: str
    page_num: int
    link_type: str
    is_hyperlinked: str
    is_valid: str
    hyperlink_points_to: str
    result: str


# ── Worker (top-level so ProcessPoolExecutor can pickle it) ──────────────────
def _extract_page_chunk(args: tuple):
    """
    Each worker opens the PDF once and processes its assigned page indices.
    Returns plain tuples for fast inter-process serialisation.
    """
    pdf_path, page_indices = args
    import fitz  # PyMuPDF — imported inside worker for subprocess safety

    tokens_raw: List[tuple] = []   # (page_idx, text, x0, y0, x1, y1)
    links_raw:  List[tuple] = []   # (page_idx, uri,  x0, y0, x1, y1)

    doc = fitz.open(pdf_path)
    for page_idx in page_indices:
        page = doc.load_page(page_idx)

        # Words with bounding boxes — 10-50x faster than pdfplumber
        for w in page.get_text("words"):
            # w = (x0, y0, x1, y1, text, block_no, line_no, word_no)
            tokens_raw.append((page_idx, w[4],
                                float(w[0]), float(w[1]),
                                float(w[2]), float(w[3])))

        # Hyperlink annotations
        for link in page.get_links():
            if link.get("kind") == fitz.LINK_URI:
                r = link["from"]   # fitz.Rect
                links_raw.append((page_idx, link["uri"],
                                   float(r.x0), float(r.y0),
                                   float(r.x1), float(r.y1)))
    doc.close()
    return tokens_raw, links_raw


# ── Parallel extraction ──────────────────────────────────────────────────────
def extract_all(pdf_path: str) -> Tuple[List[TextToken], List[LinkAnnotation]]:
    """
    Extract words + links from ALL pages.
    Pages are chunked across CPU cores for true parallelism.
    No page limit whatsoever.
    """
    import fitz
    doc        = fitz.open(pdf_path)
    n_pages    = len(doc)
    doc.close()

    cpu        = os.cpu_count() or 4
    n_workers  = min(cpu, max(1, math.ceil(n_pages / 4)))
    chunk_size = math.ceil(n_pages / n_workers)
    chunks     = [list(range(i, min(i + chunk_size, n_pages)))
                  for i in range(0, n_pages, chunk_size)]

    all_tokens: List[TextToken]      = []
    all_links:  List[LinkAnnotation] = []

    with ProcessPoolExecutor(max_workers=n_workers) as ex:
        futures = {ex.submit(_extract_page_chunk, (pdf_path, chunk)): chunk
                   for chunk in chunks}
        for fut in as_completed(futures):
            tokens_raw, links_raw = fut.result()
            for (pi, text, x0, y0, x1, y1) in tokens_raw:
                all_tokens.append(TextToken(pi, text, x0, y0, x1, y1))
            for (pi, uri, x0, y0, x1, y1) in links_raw:
                all_links.append(LinkAnnotation(pi, uri, (x0, y0, x1, y1)))

    all_tokens.sort(key=lambda t: (t.page_num, t.y0, t.x0))
    all_links.sort(key=lambda l:  l.page_num)
    return all_tokens, all_links


# ── Spatial index ────────────────────────────────────────────────────────────
BUCKET = 10

def build_spatial_index(links: List[LinkAnnotation]) -> dict:
    idx = {}
    for lnk in links:
        pg = idx.setdefault(lnk.page_num, {})
        for b in range(int(lnk.rect[1]) // BUCKET,
                       int(lnk.rect[3]) // BUCKET + 1):
            pg.setdefault(b, []).append(lnk)
    return idx


def find_link(tok: TextToken, idx: dict,
              thr: float = 0.3) -> Optional[LinkAnnotation]:
    pg = idx.get(tok.page_num)
    if not pg:
        return None
    seen = set()
    for b in range(int(tok.y0) // BUCKET, int(tok.y1) // BUCKET + 1):
        for lnk in pg.get(b, []):
            if id(lnk) in seen:
                continue
            seen.add(id(lnk))
            ix0 = max(tok.x0, lnk.rect[0]); iy0 = max(tok.y0, lnk.rect[1])
            ix1 = min(tok.x1, lnk.rect[2]); iy1 = min(tok.y1, lnk.rect[3])
            if ix1 <= ix0 or iy1 <= iy0:
                continue
            inter = (ix1 - ix0) * (iy1 - iy0)
            sa = (tok.x1 - tok.x0) * (tok.y1 - tok.y0)
            sb = (lnk.rect[2] - lnk.rect[0]) * (lnk.rect[3] - lnk.rect[1])
            m  = min(sa, sb)
            if m > 0 and inter / m >= thr:
                return lnk
    return None


# ── Address helpers ──────────────────────────────────────────────────────────
def is_mailto(u: str) -> bool:
    return u.lower().startswith("mailto:")

def is_http(u: str) -> bool:
    return u.lower().startswith(("http://", "https://", "www."))

def addr_match(displayed: str, uri: str) -> bool:
    d = displayed.strip().rstrip(".,;:!?")
    if is_mailto(uri):
        return d.lower() == uri[7:].split("?")[0].strip().lower()
    if is_http(uri):
        def n(u):
            u = u.lower().rstrip("/")
            for p in ("https://", "http://", "www."):
                if u.startswith(p): u = u[len(p):]
            return u
        return n(d) == n(uri)
    return False


# ── Line grouping & scanning ─────────────────────────────────────────────────
def group_lines(tokens: List[TextToken]) -> List[List[TextToken]]:
    if not tokens:
        return []
    lines, cur = [], [tokens[0]]
    for t in tokens[1:]:
        if t.page_num == cur[-1].page_num and abs(t.y0 - cur[-1].y0) < 4:
            cur.append(t)
        else:
            lines.append(cur); cur = [t]
    lines.append(cur)
    return lines


def scan_line(line: List[TextToken]) -> List[Tuple[str, List[TextToken]]]:
    full = " ".join(t.text for t in line)
    results = []
    for pat in (EMAIL_REGEX, URL_REGEX):
        for m in pat.finditer(full):
            addr = m.group().rstrip(".,;:!?")
            cov, pos = [], 0
            for tok in line:
                s = full.find(tok.text, pos)
                e = s + len(tok.text)
                if s < m.end() and e > m.start():
                    cov.append(tok)
                pos = s + 1
            if cov:
                results.append((addr, cov))
    return results


# ── Main validation ──────────────────────────────────────────────────────────
def validate_pdf(pdf_path: str, verbose: bool = False) -> List[ExcelRow]:
    """
    Validate every link in the PDF.
    No page limit — works for 10 pages or 10,000 pages.
    """
    print(f"📄  {pdf_path}")
    print("    Extracting pages in parallel (PyMuPDF)…")

    tokens, links = extract_all(pdf_path)

    print(f"    {len(links)} hyperlinks · {len(tokens)} tokens found")

    spatial_idx = build_spatial_index(links)
    lines       = group_lines(tokens)
    rows: List[ExcelRow] = []

    for line in lines:
        for addr, toks in scan_line(line):
            page_num  = toks[0].page_num + 1
            is_email  = bool(EMAIL_REGEX.fullmatch(addr.strip()))
            link_type = "Email" if is_email else "Web Link"

            found = [find_link(t, spatial_idx) for t in toks]
            found = [f for f in found if f]
            uris  = list({f.uri for f in found})

            is_hyp = "Yes" if uris else "No"
            h_text = uris[0] if uris else "No Link"
            valid  = "Yes"; result = "Pass"

            if not uris:
                valid = "No"; result = "Fail"
            else:
                for uri in uris:
                    if   is_email and is_http(uri):       valid = "No"; result = "Fail"
                    elif not is_email and is_mailto(uri): valid = "No"; result = "Fail"
                    elif not addr_match(addr, uri):       valid = "No"; result = "Fail"

            if verbose:
                print(f"  {'✅' if result=='Pass' else '❌'} p{page_num}: '{addr}' → {h_text}")

            rows.append(ExcelRow(addr, h_text, page_num, link_type,
                                 is_hyp, valid, h_text, result))
    return rows


# ── Excel export ─────────────────────────────────────────────────────────────
def export_excel(rows: List[ExcelRow], output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook(); ws = wb.active; ws.title = "PDF Link Validation"
    ws.append(["PDF_Link","Hyperlink_Text","Page_Numb","Link_Type",
               "Is_Hyperlinked","Is_Valid","Hyperlink_Points_To","Result"])

    hfill = PatternFill("solid", fgColor="4472C4")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for c in ws[1]:
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    pf  = PatternFill("solid", fgColor="D4EDDA"); pff = Font(color="155724", bold=True)
    nf  = PatternFill("solid", fgColor="F8D7DA"); nff = Font(color="721C24", bold=True)

    for r in rows:
        ws.append([r.pdf_link, r.hyperlink_text, r.page_num, r.link_type,
                   r.is_hyperlinked, r.is_valid, r.hyperlink_points_to, r.result])
        rc = ws[f"H{ws.max_row}"]
        rc.fill, rc.font = (pf, pff) if r.result == "Pass" else (nf, nff)

    for col, w in zip("ABCDEFGH", [30,40,12,12,14,11,40,10]):
        ws.column_dimensions[col].width = w
    for ws_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, c in enumerate(ws_row):
            c.alignment = Alignment(
                horizontal="center" if i in [2,3,4,5,7] else "left",
                vertical="center", wrap_text=True)
    wb.save(output_path)
    print(f"✅  Report saved: {output_path}\n")


# ── GUI file picker ──────────────────────────────────────────────────────────
def select_pdf_file() -> Optional[str]:
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="Select PDF File",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")])
        root.destroy(); return path
    except ImportError:
        print("❌ tkinter not available. Pass the PDF path as argument.")
        return None


# ── Entry point ──────────────────────────────────────────────────────────────
def main():
    from multiprocessing import freeze_support
    freeze_support()   # Required on Windows with ProcessPoolExecutor

    parser = argparse.ArgumentParser(
        description="Validate all hyperlinks in a PDF (unlimited pages, ultra-fast).")
    parser.add_argument("pdf", nargs="?", default=None)
    parser.add_argument("-o", "--output")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    pdf_path = args.pdf or select_pdf_file()
    if not pdf_path:
        print("❌ No file selected."); sys.exit(1)
    if not os.path.exists(pdf_path):
        print(f"❌ Not found: {pdf_path}"); sys.exit(1)

    rows = validate_pdf(pdf_path, verbose=args.verbose)

    passes = sum(1 for r in rows if r.result == "Pass")
    print("=" * 60)
    print(f"📊  Total: {len(rows)}  ✅ Pass: {passes}  ❌ Fail: {len(rows)-passes}")
    if rows: print(f"    Pass rate: {passes/len(rows)*100:.1f}%")
    print("=" * 60 + "\n")

    out = (args.output or
           os.path.splitext(os.path.basename(pdf_path))[0] + "_validation_report.xlsx")
    if not out.endswith(".xlsx"): out += ".xlsx"
    export_excel(rows, out)
    sys.exit(0 if (len(rows) - passes) == 0 else 1)


if __name__ == "__main__":
    main()
