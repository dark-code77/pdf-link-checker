"""
PDF Link Checker — Web App  (Ultra-Fast Edition v4.0)
======================================================
Speed improvements over original:
  • PyMuPDF (fitz) replaces pdfplumber  →  10–50x faster text extraction
  • Links + text in ONE pass per page   →  single file open per CPU worker
  • ProcessPoolExecutor across all cores →  true parallelism
  • Spatial Y-bucket index              →  O(1) link lookup
  • SSE live progress stream            →  browser shows real progress
  • No page limit                       →  works for any PDF size

Requirements:
    pip install flask pymupdf openpyxl

Usage:
    python app.py
    Then open: http://localhost:5050
"""

import re, os, sys, io, json, threading, webbrowser, tempfile, time, queue, math
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple

from flask import (Flask, request, jsonify, send_file,
                   render_template_string, Response, stream_with_context)

# ── Dependency check ─────────────────────────────────────────────────────────
try:
    import fitz                                  # PyMuPDF
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    DEPS_OK = True
except ImportError as e:
    DEPS_OK = False
    MISSING_DEP = str(e)

app = Flask(__name__)
UPLOAD_FOLDER = tempfile.gettempdir()

# ── Regex ─────────────────────────────────────────────────────────────────────
EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
URL_REGEX   = re.compile(r"(?:https?://|www\.)[a-zA-Z0-9\-._~:/?#\[\]@!$&'()*+,;=%]+")


# ── Data classes ──────────────────────────────────────────────────────────────
@dataclass
class LinkAnnotation:
    page_num: int; uri: str; rect: tuple

@dataclass
class TextToken:
    page_num: int; text: str
    x0: float; y0: float; x1: float; y1: float

@dataclass
class Row:
    pdf_link: str; hyperlink_text: str; page_num: int; link_type: str
    is_hyperlinked: str; is_valid: str; hyperlink_points_to: str; result: str


# ── Worker (top-level for pickling) ──────────────────────────────────────────
def _extract_page_chunk(args: tuple):
    """Open PDF once per worker; process assigned page indices."""
    pdf_path, page_indices = args
    import fitz

    tokens_raw: List[tuple] = []
    links_raw:  List[tuple] = []
    doc = fitz.open(pdf_path)
    for page_idx in page_indices:
        page = doc.load_page(page_idx)
        for w in page.get_text("words"):
            tokens_raw.append((page_idx, w[4],
                                float(w[0]), float(w[1]),
                                float(w[2]), float(w[3])))
        for link in page.get_links():
            if link.get("kind") == fitz.LINK_URI:
                r = link["from"]
                links_raw.append((page_idx, link["uri"],
                                   float(r.x0), float(r.y0),
                                   float(r.x1), float(r.y1)))
    doc.close()
    return tokens_raw, links_raw


# ── Parallel extraction ───────────────────────────────────────────────────────
def extract_all(pdf_path: str) -> Tuple[List[TextToken], List[LinkAnnotation]]:
    doc     = fitz.open(pdf_path)
    n_pages = len(doc)
    doc.close()

    cpu        = os.cpu_count() or 4
    n_workers  = min(cpu, max(1, math.ceil(n_pages / 4)))
    chunk_size = math.ceil(n_pages / n_workers)
    chunks     = [list(range(i, min(i + chunk_size, n_pages)))
                  for i in range(0, n_pages, chunk_size)]

    all_tokens: List[TextToken]      = []
    all_links:  List[LinkAnnotation] = []

    with ProcessPoolExecutor(max_workers=n_workers) as ex:
        futures = {ex.submit(_extract_page_chunk, (pdf_path, ch)): ch
                   for ch in chunks}
        for fut in as_completed(futures):
            tr, lr = fut.result()
            for (pi, text, x0, y0, x1, y1) in tr:
                all_tokens.append(TextToken(pi, text, x0, y0, x1, y1))
            for (pi, uri, x0, y0, x1, y1) in lr:
                all_links.append(LinkAnnotation(pi, uri, (x0, y0, x1, y1)))

    all_tokens.sort(key=lambda t: (t.page_num, t.y0, t.x0))
    all_links.sort(key=lambda l:  l.page_num)
    return all_tokens, all_links


# ── Spatial index ─────────────────────────────────────────────────────────────
BUCKET = 10

def build_spatial_index(links: List[LinkAnnotation]) -> dict:
    idx = {}
    for lnk in links:
        pg = idx.setdefault(lnk.page_num, {})
        for b in range(int(lnk.rect[1]) // BUCKET,
                       int(lnk.rect[3]) // BUCKET + 1):
            pg.setdefault(b, []).append(lnk)
    return idx


def find_link(tok: TextToken, idx: dict, thr: float = 0.3) -> Optional[LinkAnnotation]:
    pg = idx.get(tok.page_num)
    if not pg:
        return None
    seen = set()
    for b in range(int(tok.y0) // BUCKET, int(tok.y1) // BUCKET + 1):
        for lnk in pg.get(b, []):
            if id(lnk) in seen:
                continue
            seen.add(id(lnk))
            ix0 = max(tok.x0, lnk.rect[0]); iy0 = max(tok.y0, lnk.rect[1])
            ix1 = min(tok.x1, lnk.rect[2]); iy1 = min(tok.y1, lnk.rect[3])
            if ix1 <= ix0 or iy1 <= iy0:
                continue
            inter = (ix1 - ix0) * (iy1 - iy0)
            sa = (tok.x1 - tok.x0) * (tok.y1 - tok.y0)
            sb = (lnk.rect[2] - lnk.rect[0]) * (lnk.rect[3] - lnk.rect[1])
            m  = min(sa, sb)
            if m > 0 and inter / m >= thr:
                return lnk
    return None


# ── Address helpers ───────────────────────────────────────────────────────────
def is_mailto(u): return u.lower().startswith("mailto:")
def is_http(u):   return u.lower().startswith(("http://","https://","www."))

def addr_match(displayed, uri):
    d = displayed.strip().rstrip(".,;:!?")
    if is_mailto(uri):
        return d.lower() == uri[7:].split("?")[0].strip().lower()
    if is_http(uri):
        def n(u):
            u = u.lower().rstrip("/")
            for p in ("https://","http://","www."):
                if u.startswith(p): u = u[len(p):]
            return u
        return n(d) == n(uri)
    return False


# ── Line grouping & scanning ──────────────────────────────────────────────────
def group_lines(tokens):
    if not tokens: return []
    lines, cur = [], [tokens[0]]
    for t in tokens[1:]:
        if t.page_num == cur[-1].page_num and abs(t.y0 - cur[-1].y0) < 4:
            cur.append(t)
        else:
            lines.append(cur); cur = [t]
    lines.append(cur)
    return lines


def scan_line(line):
    full = " ".join(t.text for t in line)
    results = []
    for pat in (EMAIL_REGEX, URL_REGEX):
        for m in pat.finditer(full):
            addr = m.group().rstrip(".,;:!?")
            cov, pos = [], 0
            for tok in line:
                s = full.find(tok.text, pos); e = s + len(tok.text)
                if s < m.end() and e > m.start(): cov.append(tok)
                pos = s + 1
            if cov: results.append((addr, cov))
    return results


# ── Validate ──────────────────────────────────────────────────────────────────
def validate_pdf(pdf_path: str, progress_cb=None) -> List[Row]:
    def cb(pct, msg):
        if progress_cb: progress_cb(pct, msg)

    cb(5,  "Starting parallel extraction (PyMuPDF)…")
    tokens, links = extract_all(pdf_path)
    cb(55, f"Building index  ({len(links)} links · {len(tokens)} tokens)…")

    spatial_idx = build_spatial_index(links)
    lines       = group_lines(tokens)

    cb(62, "Validating addresses…")
    rows: List[Row] = []
    n = len(lines)
    for i, line in enumerate(lines):
        if i % max(1, n // 8) == 0:
            cb(62 + int(28 * i / max(1, n)),
               f"Scanning line {i+1} of {n}…")
        for addr, toks in scan_line(line):
            page_num  = toks[0].page_num + 1
            is_email  = bool(EMAIL_REGEX.fullmatch(addr.strip()))
            link_type = "Email" if is_email else "Web Link"

            found = [find_link(t, spatial_idx) for t in toks]
            found = [f for f in found if f]
            uris  = list({f.uri for f in found})

            is_hyp = "Yes" if uris else "No"
            h_text = uris[0] if uris else "No Link"
            valid  = "Yes"; result = "Pass"

            if not uris:
                valid = "No"; result = "Fail"
            else:
                for uri in uris:
                    if   is_email and is_http(uri):       valid = "No"; result = "Fail"
                    elif not is_email and is_mailto(uri): valid = "No"; result = "Fail"
                    elif not addr_match(addr, uri):       valid = "No"; result = "Fail"

            rows.append(Row(addr, h_text, page_num, link_type,
                            is_hyp, valid, h_text, result))

    cb(92, "Generating Excel report…")
    return rows


# ── Excel export ──────────────────────────────────────────────────────────────
def export_excel(rows, path):
    wb = Workbook(); ws = wb.active; ws.title = "PDF Link Validation"
    ws.append(["PDF_Link","Hyperlink_Text","Page_Numb","Link_Type",
               "Is_Hyperlinked","Is_Valid","Hyperlink_Points_To","Result"])
    hfill = PatternFill("solid", fgColor="1F3864")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for c in ws[1]:
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    pf  = PatternFill("solid", fgColor="C6EFCE"); pff = Font(color="276221", bold=True)
    nf  = PatternFill("solid", fgColor="FFCCCC"); nff = Font(color="9C0006", bold=True)
    for r in rows:
        ws.append([r.pdf_link, r.hyperlink_text, r.page_num, r.link_type,
                   r.is_hyperlinked, r.is_valid, r.hyperlink_points_to, r.result])
        rc = ws[f"H{ws.max_row}"]
        rc.fill, rc.font = (pf, pff) if r.result == "Pass" else (nf, nff)
    for col, w in zip("ABCDEFGH", [30,40,12,12,14,11,40,10]):
        ws.column_dimensions[col].width = w
    for ws_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, c in enumerate(ws_row):
            c.alignment = Alignment(
                horizontal="center" if i in [2,3,4,5,7] else "left",
                vertical="center", wrap_text=True)
    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# HTML TEMPLATE
# ─────────────────────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>PDF Link Checker</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet"/>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#0d0f14;--surface:#13161e;--card:#191d28;--border:#252836;
  --accent:#4f8ef7;--accent2:#f7c948;--pass:#22c55e;--fail:#ef4444;
  --text:#e8eaf0;--muted:#6b7280;--mono:'JetBrains Mono',monospace;--sans:'Syne',sans-serif;
}
body{background:var(--bg);color:var(--text);font-family:var(--sans);min-height:100vh;
  background-image:radial-gradient(ellipse 80% 60% at 50% -10%,rgba(79,142,247,.12),transparent)}
.wrap{max-width:1200px;margin:0 auto;padding:32px 24px}
header{display:flex;align-items:center;gap:16px;margin-bottom:48px}
.logo{width:44px;height:44px;background:var(--accent);border-radius:12px;display:grid;place-items:center;font-size:22px;flex-shrink:0}
h1{font-size:clamp(22px,3vw,30px);font-weight:800;letter-spacing:-.5px}
h1 span{color:var(--accent)}
.badge{margin-left:auto;font-family:var(--mono);font-size:11px;background:var(--card);
  border:1px solid var(--border);border-radius:6px;padding:4px 10px;color:var(--muted)}
.drop-zone{border:2px dashed var(--border);border-radius:20px;padding:56px 32px;
  text-align:center;cursor:pointer;transition:.25s;position:relative;overflow:hidden;background:var(--surface)}
.drop-zone::before{content:'';position:absolute;inset:0;
  background:radial-gradient(circle at 50% 0%,rgba(79,142,247,.06),transparent 70%);pointer-events:none}
.drop-zone:hover,.drop-zone.drag{border-color:var(--accent);background:rgba(79,142,247,.04)}
.drop-zone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.drop-icon{font-size:52px;margin-bottom:16px;display:block}
.drop-zone h2{font-size:20px;font-weight:700;margin-bottom:8px}
.drop-zone p{color:var(--muted);font-size:14px}
.file-name{margin-top:14px;font-family:var(--mono);font-size:13px;color:var(--accent);
  background:rgba(79,142,247,.1);border-radius:8px;padding:8px 14px;display:none}
.controls{display:flex;gap:12px;margin-top:20px;flex-wrap:wrap;align-items:center}
.ctrl-label{font-size:13px;color:var(--muted);font-weight:600;letter-spacing:.5px;text-transform:uppercase}
label.toggle{display:flex;align-items:center;gap:8px;cursor:pointer;font-size:14px;color:var(--text);user-select:none}
input[type=checkbox]{appearance:none;width:38px;height:20px;background:var(--border);
  border-radius:10px;position:relative;cursor:pointer;transition:.2s;flex-shrink:0}
input[type=checkbox]::after{content:'';position:absolute;width:14px;height:14px;
  background:#fff;border-radius:50%;top:3px;left:3px;transition:.2s}
input[type=checkbox]:checked{background:var(--accent)}
input[type=checkbox]:checked::after{left:21px}
.btn{display:inline-flex;align-items:center;gap:8px;padding:12px 28px;border-radius:12px;
  border:none;cursor:pointer;font-family:var(--sans);font-weight:700;font-size:15px;transition:.2s}
.btn-primary{background:var(--accent);color:#fff}
.btn-primary:hover{background:#3a7de8;transform:translateY(-1px)}
.btn-primary:disabled{opacity:.45;cursor:not-allowed;transform:none}
.btn-dl{background:var(--accent2);color:#0d0f14}
.btn-dl:hover{background:#e6b83a;transform:translateY(-1px)}
.progress-wrap{display:none;margin-top:28px}
.progress-bar-bg{height:8px;background:var(--border);border-radius:4px;overflow:hidden}
.progress-bar{height:100%;background:linear-gradient(90deg,var(--accent),#7eb4fc);
  border-radius:4px;width:0%;transition:.35s}
.progress-msg{font-family:var(--mono);font-size:12px;color:var(--muted);margin-top:8px}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:16px;margin:32px 0}
.stat-card{background:var(--card);border:1px solid var(--border);border-radius:16px;
  padding:20px 24px;position:relative;overflow:hidden}
.stat-card::after{content:'';position:absolute;bottom:0;left:0;right:0;height:3px}
.stat-card.total::after{background:var(--accent)}.stat-card.pass::after{background:var(--pass)}
.stat-card.fail::after{background:var(--fail)}.stat-card.rate::after{background:var(--accent2)}
.stat-card .label{font-size:11px;font-weight:700;letter-spacing:.8px;text-transform:uppercase;
  color:var(--muted);margin-bottom:8px}
.stat-card .value{font-size:36px;font-weight:800;line-height:1}
.stat-card.pass .value{color:var(--pass)}.stat-card.fail .value{color:var(--fail)}
.stat-card.rate .value{color:var(--accent2)}
.filter-bar{display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap;align-items:center}
.filter-bar span{font-size:13px;color:var(--muted);font-weight:600}
.pill{padding:6px 14px;border-radius:20px;border:1px solid var(--border);background:var(--card);
  color:var(--muted);font-size:13px;cursor:pointer;transition:.15s;font-family:var(--sans);font-weight:600}
.pill:hover,.pill.active{background:var(--accent);border-color:var(--accent);color:#fff}
.pill.pass-pill.active{background:var(--pass);border-color:var(--pass)}
.pill.fail-pill.active{background:var(--fail);border-color:var(--fail)}
.search-box{margin-left:auto;padding:7px 14px;border-radius:10px;border:1px solid var(--border);
  background:var(--card);color:var(--text);font-family:var(--mono);font-size:13px;outline:none;
  min-width:200px;transition:.2s}
.search-box:focus{border-color:var(--accent)}
.table-wrap{background:var(--card);border:1px solid var(--border);border-radius:16px;overflow:hidden}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:#111420;padding:14px 16px;text-align:left;font-size:11px;font-weight:700;
  letter-spacing:.6px;text-transform:uppercase;color:var(--muted);
  border-bottom:1px solid var(--border);white-space:nowrap}
tbody tr{border-bottom:1px solid rgba(37,40,54,.6);transition:.15s}
tbody tr:last-child{border-bottom:none}
tbody tr:hover{background:rgba(79,142,247,.04)}
td{padding:13px 16px;vertical-align:top;word-break:break-all}
.td-uri{font-family:var(--mono);font-size:11px;color:var(--muted);word-break:break-all}
.tag{display:inline-flex;align-items:center;gap:4px;padding:3px 10px;border-radius:6px;
  font-size:11px;font-weight:700;letter-spacing:.3px;white-space:nowrap}
.tag-email{background:rgba(147,51,234,.15);color:#c084fc}
.tag-web{background:rgba(79,142,247,.15);color:var(--accent)}
.tag-yes{background:rgba(34,197,94,.12);color:var(--pass)}
.tag-no{background:rgba(239,68,68,.12);color:var(--fail)}
.tag-pass{background:rgba(34,197,94,.15);color:var(--pass);font-size:12px;padding:4px 12px}
.tag-fail{background:rgba(239,68,68,.15);color:var(--fail);font-size:12px;padding:4px 12px}
.page-num{font-family:var(--mono);font-size:12px;color:var(--muted)}
.toolbar{display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap}
.toolbar-title{font-size:17px;font-weight:700;flex:1}
.row-count{font-family:var(--mono);font-size:12px;color:var(--muted)}
.empty{padding:64px;text-align:center;color:var(--muted)}
.empty span{font-size:40px;display:block;margin-bottom:12px}
.error-box{background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.25);
  border-radius:12px;padding:16px 20px;color:#fca5a5;font-family:var(--mono);
  font-size:13px;margin-top:16px;display:none}
.pagination{display:flex;justify-content:flex-end;gap:6px;margin-top:16px;align-items:center}
.pg-btn{padding:6px 12px;border-radius:8px;border:1px solid var(--border);
  background:var(--card);color:var(--muted);cursor:pointer;font-size:13px;transition:.15s}
.pg-btn:hover,.pg-btn.active{background:var(--accent);border-color:var(--accent);color:#fff}
.pg-btn:disabled{opacity:.3;cursor:not-allowed}
.pg-info{font-family:var(--mono);font-size:12px;color:var(--muted)}
@media(max-width:640px){
  .stats{grid-template-columns:1fr 1fr}
  .filter-bar{flex-direction:column;align-items:flex-start}
  .search-box{margin-left:0;width:100%}
}
</style>
</head>
<body>
<div class="wrap">

  <header>
    <div class="logo">🔗</div>
    <div>
      <h1>PDF <span>Link</span> Checker</h1>
      <p style="color:var(--muted);font-size:13px">Validate all emails &amp; URLs — unlimited pages, ultra-fast</p>
    </div>
    <div class="badge">v4.0 ⚡</div>
  </header>

  <div class="drop-zone" id="dropZone">
    <input type="file" id="fileInput" accept=".pdf"/>
    <span class="drop-icon">📄</span>
    <h2>Drop your PDF here</h2>
    <p>or click to browse — any number of pages supported</p>
    <div class="file-name" id="fileName"></div>
  </div>

  <div class="controls">
    <span class="ctrl-label">Options</span>
    <label class="toggle"><input type="checkbox" id="verboseCheck"/> Verbose output</label>
    <div style="flex:1"></div>
    <button class="btn btn-primary" id="runBtn" disabled onclick="runCheck()">▶ &nbsp;Run Check</button>
  </div>

  <div class="progress-wrap" id="progressWrap">
    <div class="progress-bar-bg"><div class="progress-bar" id="progressBar"></div></div>
    <div class="progress-msg" id="progressMsg">Starting…</div>
  </div>
  <div class="error-box" id="errorBox"></div>

  <div id="results" style="display:none">
    <div class="stats" id="statsGrid"></div>
    <div class="toolbar">
      <div class="toolbar-title">Results <span class="row-count" id="rowCount"></span></div>
      <button class="btn btn-dl" onclick="downloadExcel()">⬇&nbsp; Download Excel</button>
    </div>
    <div class="filter-bar">
      <span>Filter:</span>
      <button class="pill active" onclick="setFilter('all',this)">All</button>
      <button class="pill pass-pill" onclick="setFilter('Pass',this)">✅ Pass</button>
      <button class="pill fail-pill" onclick="setFilter('Fail',this)">❌ Fail</button>
      <input class="search-box" id="searchBox" placeholder="Search address…" oninput="applyFilters()"/>
    </div>
    <div class="table-wrap">
      <table>
        <thead><tr>
          <th>Page</th><th>PDF Link (Displayed)</th><th>Hyperlink URI</th>
          <th>Link Type</th><th>Has Link?</th><th>Valid?</th><th>Result</th>
        </tr></thead>
        <tbody id="tableBody"></tbody>
      </table>
    </div>
    <div class="pagination" id="pagination"></div>
  </div>
</div>

<script>
const PAGE_SIZE = 50;
let allRows=[], filteredRows=[], currentPage=1, activeFilter='all', lastReportPath='';

const dz=document.getElementById('dropZone'), fi=document.getElementById('fileInput');
fi.addEventListener('change',()=>handleFile(fi.files[0]));
dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('drag')});
dz.addEventListener('dragleave',()=>dz.classList.remove('drag'));
dz.addEventListener('drop',e=>{
  e.preventDefault();dz.classList.remove('drag');
  if(e.dataTransfer.files[0]) handleFile(e.dataTransfer.files[0]);
});
function handleFile(file){
  if(!file||!file.name.endsWith('.pdf')) return;
  const fn=document.getElementById('fileName');
  fn.textContent='📎  '+file.name; fn.style.display='inline-block';
  document.getElementById('runBtn').disabled=false;
  document.getElementById('results').style.display='none';
  document.getElementById('errorBox').style.display='none';
}

async function runCheck(){
  const file=fi.files[0]; if(!file) return;
  setProgress(true,3,'Uploading PDF…');
  document.getElementById('runBtn').disabled=true;
  document.getElementById('errorBox').style.display='none';
  document.getElementById('results').style.display='none';

  const fd=new FormData();
  fd.append('pdf',file);
  fd.append('verbose',document.getElementById('verboseCheck').checked);

  let ur;
  try{
    ur=await fetch('/upload',{method:'POST',body:fd});
    if(!ur.ok){showError('Upload failed');return;}
  }catch(e){showError('Upload error: '+e.message);return;}
  const {tmp_path,filename}=await ur.json();

  const sse=new EventSource('/stream?path='+encodeURIComponent(tmp_path)+'&filename='+encodeURIComponent(filename));
  sse.addEventListener('progress',e=>{const d=JSON.parse(e.data);setProgress(true,d.pct,d.msg)});
  sse.addEventListener('done',e=>{
    sse.close();
    const data=JSON.parse(e.data);
    if(data.error){showError(data.error);document.getElementById('runBtn').disabled=false;return;}
    setProgress(true,100,'Complete! ✅');
    setTimeout(()=>setProgress(false),700);
    allRows=data.rows; lastReportPath=data.report_path;
    activeFilter='all';
    document.getElementById('searchBox').value='';
    document.querySelectorAll('.pill').forEach(p=>p.classList.remove('active'));
    document.querySelectorAll('.pill')[0].classList.add('active');
    renderStats(data); applyFilters();
    document.getElementById('results').style.display='block';
    document.getElementById('runBtn').disabled=false;
  });
  sse.onerror=()=>{sse.close();showError('Connection error. Please try again.');document.getElementById('runBtn').disabled=false;};
}

function renderStats(data){
  const rate=data.total>0?((data.pass/data.total)*100).toFixed(1):'0.0';
  document.getElementById('statsGrid').innerHTML=`
    <div class="stat-card total"><div class="label">Total Found</div><div class="value">${data.total}</div></div>
    <div class="stat-card pass"><div class="label">Passed</div><div class="value">${data.pass}</div></div>
    <div class="stat-card fail"><div class="label">Failed</div><div class="value">${data.fail}</div></div>
    <div class="stat-card rate"><div class="label">Pass Rate</div><div class="value">${rate}%</div></div>`;
}

function setFilter(f,el){
  activeFilter=f;
  document.querySelectorAll('.pill').forEach(p=>p.classList.remove('active'));
  el.classList.add('active');currentPage=1;applyFilters();
}
function applyFilters(){
  const q=document.getElementById('searchBox').value.toLowerCase();
  filteredRows=allRows.filter(r=>{
    const mf=activeFilter==='all'||r.result===activeFilter;
    const ms=!q||r.pdf_link.toLowerCase().includes(q)||r.hyperlink_points_to.toLowerCase().includes(q);
    return mf&&ms;
  });
  currentPage=1;renderTable();
}

function renderTable(){
  const start=(currentPage-1)*PAGE_SIZE;
  const pageRows=filteredRows.slice(start,start+PAGE_SIZE);
  const body=document.getElementById('tableBody');
  document.getElementById('rowCount').textContent=`(${filteredRows.length} of ${allRows.length})`;
  if(!pageRows.length){
    body.innerHTML=`<tr><td colspan="7"><div class="empty"><span>🔍</span>No results match your filter.</div></td></tr>`;
    renderPagination();return;
  }
  body.innerHTML=pageRows.map(r=>`
    <tr>
      <td class="page-num">${r.page_num}</td>
      <td style="font-family:var(--mono);font-size:12px">${esc(r.pdf_link)}</td>
      <td class="td-uri">${esc(r.hyperlink_points_to)}</td>
      <td><span class="tag ${r.link_type==='Email'?'tag-email':'tag-web'}">${r.link_type==='Email'?'✉':'🌐'} ${r.link_type}</span></td>
      <td><span class="tag ${r.is_hyperlinked==='Yes'?'tag-yes':'tag-no'}">${r.is_hyperlinked==='Yes'?'✓':'✗'} ${r.is_hyperlinked}</span></td>
      <td><span class="tag ${r.is_valid==='Yes'?'tag-yes':'tag-no'}">${r.is_valid==='Yes'?'✓':'✗'} ${r.is_valid}</span></td>
      <td><span class="tag ${r.result==='Pass'?'tag-pass':'tag-fail'}">${r.result==='Pass'?'✅ Pass':'❌ Fail'}</span></td>
    </tr>`).join('');
  renderPagination();
}

function renderPagination(){
  const total=Math.ceil(filteredRows.length/PAGE_SIZE);
  const pg=document.getElementById('pagination');
  if(total<=1){pg.innerHTML='';return;}
  let html=`<span class="pg-info">Page ${currentPage} of ${total}</span>`;
  html+=`<button class="pg-btn" ${currentPage===1?'disabled':''} onclick="goPage(${currentPage-1})">‹</button>`;
  for(let i=1;i<=total;i++){
    if(i===1||i===total||Math.abs(i-currentPage)<=1)
      html+=`<button class="pg-btn ${i===currentPage?'active':''}" onclick="goPage(${i})">${i}</button>`;
    else if(Math.abs(i-currentPage)===2) html+='<span class="pg-info">…</span>';
  }
  html+=`<button class="pg-btn" ${currentPage===total?'disabled':''} onclick="goPage(${currentPage+1})">›</button>`;
  pg.innerHTML=html;
}
function goPage(p){currentPage=p;renderTable();window.scrollTo(0,0);}

async function downloadExcel(){
  if(!lastReportPath) return;
  const res=await fetch('/download?path='+encodeURIComponent(lastReportPath));
  if(!res.ok){showError('Download failed');return;}
  const blob=await res.blob();
  const a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download=lastReportPath.split('/').pop()||'report.xlsx';
  a.click();
}

function setProgress(show,pct,msg){
  document.getElementById('progressWrap').style.display=show?'block':'none';
  document.getElementById('progressBar').style.width=pct+'%';
  document.getElementById('progressMsg').textContent=msg||'';
}
function showError(msg){
  const b=document.getElementById('errorBox');
  b.textContent='❌ '+msg;b.style.display='block';setProgress(false);
}
function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
</script>
</body>
</html>"""


# ── Flask routes ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if not DEPS_OK:
        return (f"<h2>Missing dependency</h2><pre>{MISSING_DEP}</pre>"
                f"<p>Run: <code>pip install flask pymupdf openpyxl</code></p>"), 500
    return render_template_string(HTML)


@app.route("/upload", methods=["POST"])
def upload():
    if not DEPS_OK:
        return jsonify(error=f"Missing dependency: {MISSING_DEP}"), 500
    f = request.files.get("pdf")
    if not f or not f.filename.lower().endswith(".pdf"):
        return jsonify(error="Please upload a valid .pdf file"), 400
    tmp = os.path.join(UPLOAD_FOLDER, "upload_check.pdf")
    f.save(tmp)
    return jsonify(tmp_path=tmp, filename=f.filename)


@app.route("/stream")
def stream():
    """SSE: runs validation in background thread, streams progress to browser."""
    if not DEPS_OK:
        def _err():
            yield f"event: done\ndata: {json.dumps({'error': MISSING_DEP})}\n\n"
        return Response(stream_with_context(_err()), mimetype="text/event-stream")

    tmp_path = request.args.get("path", "")
    filename = request.args.get("filename", "report.pdf")

    if not tmp_path or not os.path.exists(tmp_path):
        def _err():
            yield f"event: done\ndata: {json.dumps({'error': 'File not found'})}\n\n"
        return Response(stream_with_context(_err()), mimetype="text/event-stream")

    q: queue.Queue = queue.Queue()

    def _cb(pct, msg):
        q.put(("progress", pct, msg))

    def _worker():
        try:
            rows = validate_pdf(tmp_path, progress_cb=_cb)
            base = os.path.splitext(filename)[0]
            out  = os.path.join(UPLOAD_FOLDER, f"{base}_link_report.xlsx")
            export_excel(rows, out)
            total  = len(rows)
            passed = sum(1 for r in rows if r.result == "Pass")
            q.put(("done", dict(
                total=total, passed=passed,
                **{"pass": passed, "fail": total - passed},
                report_path=out,
                rows=[{
                    "page_num":            r.page_num,
                    "pdf_link":            r.pdf_link,
                    "hyperlink_text":      r.hyperlink_text,
                    "link_type":           r.link_type,
                    "is_hyperlinked":      r.is_hyperlinked,
                    "is_valid":            r.is_valid,
                    "hyperlink_points_to": r.hyperlink_points_to,
                    "result":              r.result,
                } for r in rows]
            )))
        except Exception as e:
            q.put(("error", str(e)))

    threading.Thread(target=_worker, daemon=True).start()

    def _generate():
        while True:
            item = q.get()
            if item[0] == "progress":
                yield f"event: progress\ndata: {json.dumps({'pct':item[1],'msg':item[2]})}\n\n"
            elif item[0] == "done":
                yield f"event: done\ndata: {json.dumps(item[1])}\n\n"
                break
            elif item[0] == "error":
                yield f"event: done\ndata: {json.dumps({'error':item[1]})}\n\n"
                break

    return Response(stream_with_context(_generate()), mimetype="text/event-stream")


@app.route("/download")
def download():
    path = request.args.get("path", "")
    if not path or not os.path.exists(path):
        return "File not found", 404
    return send_file(path, as_attachment=True,
                     download_name=os.path.basename(path),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Launch ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    from multiprocessing import freeze_support
    freeze_support()   # Required on Windows

    PORT = 5050
    url  = f"http://localhost:{PORT}"
    print(f"\n{'─'*54}")
    print(f"  🔗  PDF Link Checker  —  v4.0 Ultra-Fast ⚡")
    print(f"{'─'*54}")
    print(f"  Open in browser : {url}")
    print(f"  Engine          : PyMuPDF (10–50× faster)")
    print(f"  CPU cores used  : {os.cpu_count() or 4}")
    print(f"  Press Ctrl+C to stop.")
    print(f"{'─'*54}\n")

    def _open():
        time.sleep(1.2); webbrowser.open(url)
    threading.Thread(target=_open, daemon=True).start()

    app.run(port=PORT, debug=False, threaded=True)
